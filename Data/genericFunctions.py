from Data import variables
from datetime import datetime
import openpyxl
import random


def capture_screenshot():
    variables.driver.get_screenshot_as_png()
    now = datetime.now()
    current_time = now.strftime("%H.%M.%S")
    current_date = now.date()
    abc = variables.screenshot + str(current_date) + ' ' + str(current_time) + '.png'
    # print(abc)
    variables.driver.save_screenshot(abc)


def random_date(x):
    num = random.randrange(0, x)
    vallue = str(num)
    return vallue


def getValueFromExcel():
    sourceBook = openpyxl.load_workbook(variables.source_excel)
    source_sheet = sourceBook.get_sheet_by_name("Test Cases")
    max_row = source_sheet.max_row
    # print(max_row)
    for i in range(2, max_row + 1):
        data_value = source_sheet.cell(i, 1).value
        # print(data_value)
        return data_value


def getFlagValue(testcaseid):
    if len(testcaseid) != 0:
        sourceBook = openpyxl.load_workbook(variables.source_excel)
        source_sheet = sourceBook.get_sheet_by_name("Test Cases")
        max_row = source_sheet.max_row
        for i in range(2, max_row + 1):
            flag = source_sheet.cell(i, 1).value
            tc_id = source_sheet.cell(i, 2).value
            if len(tc_id) != 0:
                if tc_id == testcaseid:
                    if flag == 'Y':
                        return_state = "The test case with test ID : ", tc_id, "is to be executed as it has the flag '", flag, "'"
                        print(return_state)
                    else:
                        return_state = "The test case with test ID : ", tc_id, "is NOT to be executed as it has the flag '", flag, "'"
                        print(return_state)
                        break
                else:
                    print("error")
    else:
        return_state = "Test case ID should be of at least one character"
        return return_state
